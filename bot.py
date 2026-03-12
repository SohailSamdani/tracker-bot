import os
import re
import json
import logging
import httpx
import asyncio
import base64
from datetime import datetime
from telegram import Update
from telegram.ext import ApplicationBuilder, CommandHandler, MessageHandler, filters, ContextTypes
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(format='%(asctime)s - %(levelname)s - %(message)s', level=logging.INFO)
logger = logging.getLogger(__name__)

TOKEN = os.environ.get("BOT_TOKEN", "")
GEMINI_KEY = os.environ.get("GEMINI_API_KEY", "")
EXCEL_FILE = "tracker_log.xlsx"

# ── Gemini Vision ─────────────────────────────────────────────────────────────

async def extract_nodes_with_gemini(image_bytes: bytes) -> list:
    b64 = base64.standard_b64encode(image_bytes).decode("utf-8")

    if image_bytes[:8] == b'\x89PNG\r\n\x1a\n':
        media_type = "image/png"
    elif image_bytes[:2] == b'\xff\xd8':
        media_type = "image/jpeg"
    else:
        media_type = "image/jpeg"

    prompt = """This is a screenshot from GeniusVision solar tracker software.
Extract all rows from the table and return ONLY a JSON array like this:
[{"node_id": "30C8", "master": "Master 20"}, ...]

Rules:
- node_id is the 4-character hex value in the Node ID column
- master is the value in the Master column e.g. "Master 20"
- Return ONLY the JSON array, no explanation, no markdown backticks"""

    payload = {
        "contents": [
            {
                "parts": [
                    {
                        "inline_data": {
                            "mime_type": media_type,
                            "data": b64
                        }
                    },
                    {
                        "text": prompt
                    }
                ]
            }
        ],
        "generationConfig": {
            "temperature": 0,
            "maxOutputTokens": 2000
        }
    }

    url = f"https://generativelanguage.googleapis.com/v1beta/models/gemini-2.0-flash:generateContent?key={GEMINI_KEY}"

    async with httpx.AsyncClient(timeout=60) as client:
        for attempt in range(3):
            response = await client.post(url, json=payload)
            if response.status_code == 429:
                await asyncio.sleep(10)
                continue
            response.raise_for_status()
            break

        data = response.json()
        logger.info(f"Full Gemini response: {json.dumps(data)[:500]}")

        if "candidates" not in data:
            raise ValueError(f"Unexpected response: {json.dumps(data)[:300]}")

        candidate = data["candidates"][0]
        if "content" not in candidate:
            raise ValueError(f"No content in candidate: {json.dumps(candidate)[:300]}")

        text = candidate["content"]["parts"][0]["text"].strip()
        logger.info(f"Gemini text: {text[:300]}")

        text = re.sub(r"```json|```", "", text).strip()
        nodes_raw = json.loads(text)
        return [(n["node_id"].strip().upper(), n["master"].strip()) for n in nodes_raw]


# ── Excel helpers ─────────────────────────────────────────────────────────────

def get_or_create_workbook():
    if os.path.exists(EXCEL_FILE):
        return openpyxl.load_workbook(EXCEL_FILE)
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Tracker Log"
    headers = ["S.No", "Node ID", "Master", "Manual Cmd Angle", "Date Added", "Source"]
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(name="Arial", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill("solid", start_color="1F4E79")
    center = Alignment(horizontal="center", vertical="center")
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center
        cell.border = border
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 20
    ws.column_dimensions["E"].width = 16
    ws.column_dimensions["F"].width = 22
    ws.freeze_panes = "A2"
    wb.save(EXCEL_FILE)
    return wb


def get_existing_node_ids(wb):
    ws = wb["Tracker Log"]
    existing = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[1]:
            existing.add(str(row[1]).strip().upper())
    return existing


def append_nodes(nodes, source_name):
    wb = get_or_create_workbook()
    ws = wb["Tracker Log"]
    existing = get_existing_node_ids(wb)
    thin = Side(style="thin", color="BFBFBF")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    cell_font = Font(name="Arial", size=10)
    alt_fill = PatternFill("solid", start_color="DCE6F1")

    added = 0
    skipped = 0
    today = datetime.now().strftime("%d-%m-%Y")

    for node_id, master in nodes:
        node_id_upper = node_id.strip().upper()
        if node_id_upper in existing:
            skipped += 1
            continue
        row_num = ws.max_row + 1
        s_no = row_num - 1
        row_data = [s_no, node_id_upper, master, "", today, source_name]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.font = cell_font
            cell.alignment = center
            cell.border = border
            if s_no % 2 == 0:
                cell.fill = alt_fill
        existing.add(node_id_upper)
        added += 1

    wb.save(EXCEL_FILE)
    return added, skipped


# ── Bot handlers ──────────────────────────────────────────────────────────────

async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg = (
        "Tracker Node Bot\n\n"
        "Send me GeniusVision screenshots and I will extract Node IDs and Masters into Excel.\n\n"
        "Commands:\n"
        "/export — Download the current Excel file\n"
        "/count  — See total nodes per Master\n"
        "/clear  — Delete all data and start fresh\n"
        "/help   — Show this message"
    )
    await update.message.reply_text(msg)


async def handle_photo(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Reading image, please wait...")
    try:
        photo = update.message.photo[-1]
        file = await context.bot.get_file(photo.file_id)
        image_bytes = await file.download_as_bytearray()

        nodes = await extract_nodes_with_gemini(bytes(image_bytes))

        if not nodes:
            await update.message.reply_text(
                "Could not extract any Node IDs from this image.\n"
                "Make sure the screenshot clearly shows the Node ID and Master columns."
            )
            return

        source_name = f"photo_{datetime.now().strftime('%d%m%Y_%H%M%S')}"
        added, skipped = append_nodes(nodes, source_name)

        reply = (
            f"Extracted {len(nodes)} nodes from image.\n"
            f"New entries added: {added}\n"
            f"Duplicates skipped: {skipped}\n\n"
            f"Use /export to download the Excel file."
        )
        await update.message.reply_text(reply)

    except Exception as e:
        logger.error(f"Error: {e}")
        await update.message.reply_text(f"Error processing image: {str(e)}")


async def export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not os.path.exists(EXCEL_FILE):
        await update.message.reply_text("No data yet. Send some screenshots first.")
        return
    wb = get_or_create_workbook()
    ws = wb["Tracker Log"]
    count = ws.max_row - 1
    with open(EXCEL_FILE, "rb") as f:
        filename = f"Tracker_Log_{datetime.now().strftime('%d%m%Y_%H%M')}.xlsx"
        await update.message.reply_document(
            document=f,
            filename=filename,
            caption=f"Current log — {count} nodes total."
        )


async def count(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not os.path.exists(EXCEL_FILE):
        await update.message.reply_text("No data yet.")
        return
    wb = get_or_create_workbook()
    ws = wb["Tracker Log"]
    total = ws.max_row - 1
    masters = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[2]:
            masters[row[2]] = masters.get(row[2], 0) + 1
    lines = [f"Total nodes: {total}\n"]
    for m in sorted(masters.keys(), key=lambda x: int(x.split()[-1]) if x.split()[-1].isdigit() else 0):
        lines.append(f"  {m}: {masters[m]}")
    await update.message.reply_text("\n".join(lines))


async def clear(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if os.path.exists(EXCEL_FILE):
        os.remove(EXCEL_FILE)
    get_or_create_workbook()
    await update.message.reply_text("All data cleared. Ready for fresh data.")


async def help_cmd(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await start(update, context)


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    if not TOKEN:
        raise ValueError("BOT_TOKEN environment variable not set.")
    if not GEMINI_KEY:
        raise ValueError("GEMINI_API_KEY environment variable not set.")
    app = ApplicationBuilder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("export", export))
    app.add_handler(CommandHandler("count", count))
    app.add_handler(CommandHandler("clear", clear))
    app.add_handler(CommandHandler("help", help_cmd))
    app.add_handler(MessageHandler(filters.PHOTO, handle_photo))
    logger.info("Bot started.")
    app.run_polling()


if __name__ == "__main__":
    main()
